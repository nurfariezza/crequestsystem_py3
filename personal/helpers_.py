import  os, shutil
from .import models
from os import listdir
from os.path import isfile, join
from django.conf import settings
from django.http.response import HttpResponse, HttpResponseServerError
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import StringIO
from datetime import datetime, timedelta


def queryexcellist(req):
    b = None
    k = None

    try:
        q, x = getsearchquery(req)
        l = q.all().order_by('-reqdate')
        wb = Workbook()
        ws = wb.active
        k = StringIO()

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 41
        ws.column_dimensions['D'].width = 20

        lh = ['Request ID', 'Application', 'Creator', 'Request DateTime', 'Description', 'Head Remark', 'Remark', 'Status']
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        for o in l:
            ws.cell(row=i, column=1).value = o.id
            ws.cell(row=i, column=2).value = o.app.name
            ws.cell(row=i, column=3).value = '{} ({})'.format(o.requser.username, o.requser.email)
            ws.cell(row=i, column=4).value = o.reqdate
            ws.cell(row=i, column=5).value = o.description
            ws.cell(row=i, column=6).value = o.description2
            ws.cell(row=i, column=7).value = o.remark
            ws.cell(row=i, column=8).value = o.reqstatusstr

            i += 1

        wb.save(k)
        b = k.getvalue()

    finally:
        if k is not None:
            k.close()

    return b

def getsearchquery(req):

    q = models.ReqForm.objects
    q = q.exclude(reqstatus=9)
    
    d = req.session.get('search')
    
    if req.user.is_approver is True:
        q = models.ReqForm.objects
        q = q.exclude(reqstatus=9)

        d = req.session.get('search')
        
    elif req.user.dept =="rnd":
        q = models.ReqForm.objects
        q = q.exclude(reqstatus=9)

        d = req.session.get('search')
       
    elif req.user.dept == req.user.dept:
        #print (req.user.id)
        y = models.User.objects.filter(dept=req.user.dept).values('id')
        q = models.ReqForm.objects.filter(requser__id__in = y)


   
    else:
        
        qc =  (Q(requser__id= req.user.id))
        q = q.filter(qc)
        

    #print req.user.is_approver
    if d is None:
        return q, d 

    o = models.Search()
    o.setfromsessiondic(d)

    if o.keyword is not None:
        qk = (Q(pk=tryint(o.keyword)) | 
              Q(app__name__icontains=o.keyword) | 
              Q(description__icontains=o.keyword) | 
              Q(remark__icontains=o.keyword) | 
              Q(requser__username__icontains=o.keyword) | 
              Q(requser__email__icontains=o.keyword))


    
        if o.statuslist is not None:
            qs = Q(reqstatus__in=o.statuslist)
            #dx = Q(reqstatus__not__in=9)

            q = q.filter(qs)#, qc

        else:
            q = q.filter(qk)


    else:
        if o.statuslist is not None:
            qs = Q(reqstatus__in=o.statuslist)
            q = q.filter(qs)

    return q, o

def countlist(x):
    return models.ReqForm.objects.filter(reqstatus=x).count()

def ensure_dir(f):
    if not os.path.exists(f):
        os.makedirs(f)

def upload_file(req, path):
    f = req.FILES['file']
    ensure_dir(path)

    filename = join(path, f.name)
    with open(filename, 'wb+') as dest:
        for chunk in f.chunks():
            dest.write(chunk)

    return f.name

def movefile(o, req):
    userdir = join(settings.TMP_UPLOAD_PATH, getuserdir(req))

    if not os.path.exists(userdir):
        return
        
    dest = join(settings.UPLOAD_PATH, getreqdir(o.id))
    shutil.move(userdir, dest)

    path = dest
    filelist = [f for f in listdir(path) if isfile(join(path, f))]

    for f in filelist:
        x = models.ReqFile()
        x.reqform = o
        x.filename = f
        x.uploaduser = o.requser
        x.save()

def tryint(i):
    x = 0

    try:
        if i.isdigit():
            x = int(i)

    except:
        pass

    return x

def sendfile(b, filename, content_type):
    r = HttpResponse(b, content_type=content_type)
    r['Content-Disposition'] = 'attachment; filename="{0}"'.format(filename)
    return r

def getuserdir(req):
    return '__{0}__'.format(req.user.id)

def getreqdir(i):
    return '__{0}__'.format(i)


def getsearchquerytemp(req):
    
    q = models.ReqForm.objects
    
    d = req.session.get('search')
    
    if req.user.is_approver is True:
        q = models.ReqForm.objects

        d = req.session.get('search')
        
    elif req.user.dept =="rnd":
        q = models.ReqForm.objects

        d = req.session.get('search')
       
    elif req.user.dept == req.user.dept:
        #print (req.user.id)
        y = models.User.objects.filter(dept=req.user.dept).values('id')
        q = models.ReqForm.objects.filter(requser__id__in = y)


   
    else:
        
        qc =  (Q(requser__id= req.user.id))
        q = q.filter(qc)
    #print req.user.is_approver
    if d is None:
        return q, d 

    o = models.Search()
    o.setfromsessiondic(d)

    if o.keyword is not None:
        qk = (Q(pk=tryint(o.keyword)) | 
              Q(app__name__icontains=o.keyword) | 
              Q(description__icontains=o.keyword) | 
              Q(remark__icontains=o.keyword) | 
              Q(requser__username__icontains=o.keyword) | 
              Q(requser__email__icontains=o.keyword))


    
        if o.statuslist is not None:
            qs = Q(reqstatus__in=o.statuslist)
            q = q.filter(qs, q)#, qc

        else:
            q = q.filter(qk)


    else:
        if o.statuslist is not None:
            qs = Q(reqstatus__in=o.statuslist)
            q = q.filter(qs)

    return q, o

def verify_filter(req):
    #start = datetime(2017, 6, 1)
    start = datetime.today()
   
    end_month = datetime.today() - timedelta(days=90)
    print(start)
    b = models.ReqForm.objects.filter(reqdate__gte=end_month).filter(reqdate__lte=start).filter(reqstatus=6).filter(verified=0)
    #b = b.exclude(reqstatus=9)

    return b


